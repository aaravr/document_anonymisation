"""Apply replacements to an XLSX workbook via openpyxl.

We open the workbook (preserving formatting) and overwrite the cells flagged
by the plan with the post-edit text. Cells not flagged are left untouched.
"""
from __future__ import annotations

from idp_anonymiser.agent.state import (
    AnonymisationPlan,
    ResolvedEntity,
)
from idp_anonymiser.document.layout_model import ExtractedDocument


def _build_cell_edits(
    extracted: ExtractedDocument,
    plan: AnonymisationPlan,
    resolved: list[ResolvedEntity],
) -> dict[tuple[str, str], str]:
    """(sheet_name, coordinate) -> new_text for each cell that needs editing."""
    by_entity = {r.entity_id: r for r in plan.replacements}
    block_index = {b.block_id: b for b in extracted.blocks if b.block_id}

    # collect per-block list of edits
    per_block: dict[str, list[tuple[int, int, str]]] = {}
    for ent in resolved:
        rep = by_entity.get(ent.entity_id)
        if rep is None:
            continue
        for det in ent.detections:
            if det.span.start is None or det.span.end is None:
                continue
            for bid, block in block_index.items():
                if not bid.startswith("xlsx:"):
                    continue
                if block.start <= det.span.start and det.span.end <= block.end:
                    local_s = det.span.start - block.start
                    local_e = det.span.end - block.start
                    per_block.setdefault(bid, []).append(
                        (local_s, local_e, rep.replacement_value)
                    )
                    break

    out: dict[tuple[str, str], str] = {}
    for bid, edits in per_block.items():
        block = block_index[bid]
        sheet = block.metadata.get("sheet_name")
        coord = block.metadata.get("coordinate")
        if sheet is None or coord is None:
            continue
        new_text = block.text
        for s, e, repl in sorted(edits, key=lambda x: x[0], reverse=True):
            new_text = new_text[:s] + repl + new_text[e:]
        out[(sheet, coord)] = new_text
    return out


def rewrite_xlsx(
    input_path: str,
    output_path: str,
    extracted: ExtractedDocument,
    plan: AnonymisationPlan,
    resolved: list[ResolvedEntity],
) -> int:
    from openpyxl import load_workbook

    cell_edits = _build_cell_edits(extracted, plan, resolved)
    wb = load_workbook(filename=input_path)
    applied = 0
    for (sheet_name, coord), new_value in cell_edits.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws[coord].value = new_value
        applied += 1
    wb.save(output_path)
    wb.close()
    return applied


__all__ = ["rewrite_xlsx"]
