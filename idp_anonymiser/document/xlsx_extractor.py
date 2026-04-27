"""Extract text from XLSX workbooks via openpyxl.

We iterate every sheet/cell and capture string values. Numeric and date cells
are kept structurally (so the rewriter can re-emit them) but excluded from
detection unless they look like sensitive identifiers.
"""
from __future__ import annotations

from typing import Any

from idp_anonymiser.document.layout_model import (
    ExtractedDocument,
    ExtractedTextBlock,
    XlsxCell,
)
from idp_anonymiser.document.loader import LoadedDocument


def extract_xlsx(loaded: LoadedDocument) -> ExtractedDocument:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(loaded.path), data_only=False, read_only=False)
    parts: list[str] = []
    blocks: list[ExtractedTextBlock] = []
    cells: list[XlsxCell] = []
    cursor = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value: Any = cell.value
                if value is None:
                    continue
                # Only stringify what makes textual sense; preserve original
                # numeric/datetime via the XlsxCell record.
                str_val = str(value)
                cells.append(
                    XlsxCell(
                        sheet_name=sheet_name,
                        row=cell.row,
                        column=cell.column,
                        value=value,
                        coordinate=cell.coordinate,
                    )
                )
                if not str_val.strip():
                    continue
                start = cursor
                parts.append(str_val)
                cursor += len(str_val)
                end = cursor
                parts.append("\n")
                cursor += 1
                blocks.append(
                    ExtractedTextBlock(
                        text=str_val,
                        start=start,
                        end=end,
                        block_id=f"xlsx:{sheet_name}:{cell.coordinate}",
                        metadata={
                            "sheet_name": sheet_name,
                            "row": cell.row,
                            "column": cell.column,
                            "coordinate": cell.coordinate,
                        },
                    )
                )
    wb.close()
    return ExtractedDocument(
        flat_text="".join(parts),
        blocks=blocks,
        xlsx_cells=cells,
        metadata={"format": "xlsx", "sheets": list(wb.sheetnames) if hasattr(wb, "sheetnames") else []},
    )
